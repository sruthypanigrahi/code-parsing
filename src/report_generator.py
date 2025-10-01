"""Minimal report generator with OOP principles."""

import json
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Any, Union

import openpyxl


class BaseReportGenerator(ABC):  # Abstraction
    def __init__(self, output_dir: Path):
        self._output_dir = output_dir  # Encapsulation
        self._output_dir.mkdir(parents=True, exist_ok=True)

    @abstractmethod  # Abstraction
    def generate(self, data: dict[str, Any]) -> Path:
        pass


class JSONReportGenerator(BaseReportGenerator):  # Inheritance
    def generate(self, data: dict[str, Any]) -> Path:  # Polymorphism
        report_file = self._output_dir / "parsing_report.json"
        try:
            report: dict[str, Any] = {
                "metadata": {
                    "title": "USB PD Report",
                    "generated": datetime.now().isoformat(),
                },
                "summary": data,
                "validation": {
                    "status": "PASS" if data.get("content_items", 0) > 1000 else "FAIL"
                },
            }
        except (TypeError, ValueError) as e:
            raise RuntimeError(f"Cannot create report data: {e}") from e
        try:
            with open(report_file, "w") as f:
                json.dump(report, f, indent=2)
        except (PermissionError, OSError) as e:
            raise RuntimeError(f"Cannot write JSON report: {e}") from e
        return report_file


class ExcelReportGenerator(BaseReportGenerator):  # Inheritance
    def generate(self, data: dict[str, Any]) -> Path:  # Polymorphism
        excel_file = self._output_dir / "validation_report.xlsx"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
        except Exception as e:
            raise RuntimeError(f"Cannot create Excel workbook: {e}") from e
        try:
            ws.title = "Validation"  # type: ignore

            # Headers
            ws["A1"] = "Metric"  # type: ignore
            ws["B1"] = "Value"  # type: ignore

            # Data
            metrics: list[tuple[str, Union[int, str]]] = [
                ("Pages", data.get("pages", 0)),
                ("Content Items", data.get("content_items", 0)),
                ("TOC Entries", data.get("toc_entries", 0)),
                ("Status", "PASS" if data.get("content_items", 0) > 1000 else "FAIL"),
            ]

            for i, (metric, value) in enumerate(metrics, 2):
                ws[f"A{i}"] = metric  # type: ignore
                ws[f"B{i}"] = value  # type: ignore
        except Exception as e:
            raise RuntimeError(f"Cannot write Excel data: {e}") from e

        try:
            wb.save(excel_file)
        except (PermissionError, OSError) as e:
            raise RuntimeError(f"Cannot save Excel file: {e}") from e
        return excel_file


class ReportFactory:  # Factory pattern
    _ALLOWED_TYPES = {"json", "excel"}  # Authorized report types

    @staticmethod
    def create_generator(report_type: str, output_dir: Path) -> BaseReportGenerator:
        # Validate and sanitize input
        if not report_type.strip():
            raise ValueError("Report type must be a non-empty string")

        clean_type = report_type.strip().lower()
        if clean_type not in ReportFactory._ALLOWED_TYPES:
            raise ValueError(f"Unauthorized report type: {report_type}")

        if clean_type == "json":
            return JSONReportGenerator(output_dir)  # Polymorphism
        elif clean_type == "excel":
            return ExcelReportGenerator(output_dir)  # Polymorphism
        raise ValueError(f"Invalid report type: {report_type}")
