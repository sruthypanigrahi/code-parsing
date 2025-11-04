"""Validation report generators with OOP principles."""

from __future__ import annotations

import json
import logging
from abc import ABC
from pathlib import Path
from typing import Any

from src.config.constants import MIN_CONTENT_THRESHOLD
from src.utils.mixins import OutputDirMixin

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None  # type: ignore
    Font = None  # type: ignore

HAS_OPENPYXL = openpyxl is not None


class BaseValidator(OutputDirMixin, ABC):
    """Template method base for validation report generators."""

    def __init__(self, output_dir: Path) -> None:
        super().__init__(output_dir)
        self.__validation_results: dict[str, Any] = {}

    @property
    def validation_results(self) -> dict[str, Any]:
        """Return a snapshot of the most recent validation results."""
        return self.__validation_results.copy()

    def generate_validation(
        self, toc_data: list[Any], spec_data: list[Any]
    ) -> Path:
        """Run the validation pipeline using a template method."""
        results = self._calculate_results(toc_data, spec_data)
        output_path = self._write_report(toc_data, spec_data, results)
        self._store_results(results)
        return output_path

    def _calculate_results(
        self, toc_data: list[Any], spec_data: list[Any]
    ) -> dict[str, Any]:
        """Derive shared validation metrics."""
        status = "PASS" if len(spec_data) > MIN_CONTENT_THRESHOLD else "FAIL"
        return {
            "toc_entries": len(toc_data),
            "content_items": len(spec_data),
            "status": status,
        }

    def _write_report(
        self,
        toc_data: list[Any],
        spec_data: list[Any],
        results: dict[str, Any],
    ) -> Path:
        """Persist validation results to disk."""
        msg = f"{self.__class__.__name__} must implement _write_report()."
        raise NotImplementedError(msg)

    def _store_results(self, results: dict[str, Any]) -> None:
        """Cache the results for later inspection."""
        self.__validation_results = results.copy()


class XLSValidator(BaseValidator):
    """Excel-based validation report generator."""

    def _write_report(
        self,
        toc_data: list[Any],
        spec_data: list[Any],
        results: dict[str, Any],
    ) -> Path:
        if not HAS_OPENPYXL or openpyxl is None:
            raise ImportError("openpyxl required for Excel reports")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"  # type: ignore[arg-type]

        ws["A1"] = "USB PD Validation Report"  # type: ignore[index]
        ws["A1"].font = Font(bold=True, size=14)  # type: ignore[assignment]

        metrics: list[tuple[str, int | str]] = [
            ("TOC Entries", results["toc_entries"]),
            ("Content Items", results["content_items"]),
            ("Status", results["status"]),
        ]

        for i, (metric, value) in enumerate(metrics, 3):
            ws[f"A{i}"] = metric  # type: ignore[index]
            ws[f"B{i}"] = value  # type: ignore[index]

        xlsx_file = self.output_dir / "validation_report.xlsx"
        wb.save(xlsx_file)
        return xlsx_file


class JSONValidator(BaseValidator):
    """JSON-based validation report generator."""

    def _write_report(
        self,
        toc_data: list[Any],
        spec_data: list[Any],
        results: dict[str, Any],
    ) -> Path:
        report: dict[str, Any] = {
            "validation_report": {
                **results,
                "timestamp": "2024-01-01T00:00:00Z",
            }
        }
        json_file = self.output_dir / "validation_report.json"
        with open(json_file, "w", encoding="utf-8") as handle:
            json.dump(report, handle, indent=2)
        return json_file


class ValidationGeneratorFactory:
    """Factory for validation generators."""

    __VALIDATORS: dict[str, type[BaseValidator]] = {
        "excel": XLSValidator,
        "json": JSONValidator,
    }

    @classmethod
    def create(cls, validator_type: str, output_dir: Path) -> BaseValidator:
        """Create a validator instance of the requested type."""
        key = validator_type.lower()
        if key not in cls.__VALIDATORS:
            raise ValueError(f"Unknown validator type: {validator_type}")
        if key == "excel" and not HAS_OPENPYXL:
            raise RuntimeError("Excel validator requested but openpyxl missing")
        return cls.__VALIDATORS[key](output_dir)


class ValidationStrategySelector:
    """Selects the most appropriate validation generator."""

    _PRIORITY_ORDER: tuple[str, ...] = ("excel", "json")

    @classmethod
    def select(
        cls, output_dir: Path, preferred: str | None = None
    ) -> BaseValidator:
        """Return a validator instance using priority and availability."""
        candidates: list[str] = []
        if preferred:
            candidates.append(preferred.lower())
        candidates.extend(
            validator_type
            for validator_type in cls._PRIORITY_ORDER
            if validator_type not in candidates
        )

        for candidate in candidates:
            try:
                return ValidationGeneratorFactory.create(candidate, output_dir)
            except (RuntimeError, ValueError):
                continue
        raise RuntimeError("No validation generators available")


class ValidationDataLoader:
    """Encapsulates loading of JSONL validation inputs."""

    def __init__(self, logger: logging.Logger | None = None) -> None:
        self._logger = logger or logging.getLogger(__name__)

    def load_jsonl(self, source: Path) -> list[Any]:
        """Load structured rows from a JSONL file."""
        if not source.exists():
            return []

        entries: list[Any] = []
        try:
            with open(source, encoding="utf-8") as handle:
                for line in handle:
                    if not line.strip():
                        continue
                    try:
                        entries.append(json.loads(line))
                    except json.JSONDecodeError as exc:
                        self._logger.debug("Skipping malformed JSON line: %s", exc)
        except OSError as exc:
            self._logger.debug("Failed to load %s: %s", source, exc)
        return entries


class ValidationReportService:
    """Coordinates data loading and validator execution."""

    def __init__(
        self,
        output_dir: Path,
        loader: ValidationDataLoader | None = None,
    ) -> None:
        self._output_dir = output_dir
        self._loader = loader or ValidationDataLoader()

    def generate(
        self,
        toc_file: Path,
        spec_file: Path,
        validator_type: str | None = None,
    ) -> Path:
        """Generate a validation report using the configured strategy."""
        toc_data = self._loader.load_jsonl(toc_file)
        spec_data = self._loader.load_jsonl(spec_file)
        validator = ValidationStrategySelector.select(
            self._output_dir, validator_type
        )
        return validator.generate_validation(toc_data, spec_data)


def create_validation_report(
    output_dir: Path,
    toc_file: Path,
    spec_file: Path,
    validator_type: str | None = None,
) -> Path:
    """Backward-compatible helper to generate a validation report."""
    service = ValidationReportService(output_dir)
    return service.generate(toc_file, spec_file, validator_type)
