"""Validation report generator with OOP principles."""

import json
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font


class BaseValidator(ABC):  # Abstraction
    def __init__(self, output_dir: Path):
        self._output_dir = output_dir  # Encapsulation

    @abstractmethod  # Abstraction
    def generate_validation(self, toc_data: list[Any], spec_data: list[Any]) -> Path:
        pass


class XLSValidator(BaseValidator):  # Inheritance
    def generate_validation(
        self, toc_data: list[Any], spec_data: list[Any]
    ) -> Path:  # Polymorphism
        xlsx_file = self._output_dir / "validation_report.xlsx"
        wb = openpyxl.Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"  # type: ignore
        self._create_summary_sheet(ws_summary, toc_data, spec_data)

        # TOC vs Content comparison
        ws_comparison = wb.create_sheet("TOC_vs_Content")
        self._create_comparison_sheet(ws_comparison, toc_data, spec_data)

        try:
            wb.save(xlsx_file)
        except (PermissionError, OSError) as e:
            raise RuntimeError(f"Cannot save validation report: {e}") from e
        return xlsx_file

    def _create_summary_sheet(
        self, ws: Any, toc_data: list[Any], spec_data: list[Any]
    ) -> None:  # Encapsulation
        # Headers
        ws["A1"] = "USB PD Specification Validation Report"  # type: ignore
        ws["A1"].font = Font(bold=True, size=14)  # type: ignore

        # Metrics
        metrics: list[tuple[str, int | float | str]] = [
            ("TOC Entries", len(toc_data)),
            ("Content Items", len(spec_data)),
            ("Coverage %", round(len(spec_data) / 1046 * 100, 1)),
            ("Status", "PASS" if len(spec_data) > 1000 else "FAIL"),
        ]

        for i, (metric, value) in enumerate(metrics, 3):
            ws[f"A{i}"] = metric  # type: ignore
            ws[f"B{i}"] = value  # type: ignore

    def _create_comparison_sheet(
        self, ws: Any, toc_data: list[Any], spec_data: list[Any]
    ) -> None:  # Encapsulation
        # Headers
        headers = ["Section", "TOC Title", "Content Found", "Status"]
        for i, header in enumerate(headers, 1):
            ws.cell(1, i, header).font = Font(bold=True)  # type: ignore

        # Data rows
        for i, toc_item in enumerate(toc_data[:50], 2):  # Limit to 50
            section_id = getattr(toc_item, "section_id", f"S{i}")
            title = getattr(toc_item, "title", "Unknown")

            # Check if content exists for this TOC item
            content_found = any(
                item.get("section_id", "").startswith(section_id[:3])
                for item in spec_data[:100]  # Sample check
            )

            ws.cell(i, 1, section_id)  # type: ignore
            ws.cell(i, 2, title)  # type: ignore
            ws.cell(i, 3, "Yes" if content_found else "No")  # type: ignore
            ws.cell(i, 4, "✓" if content_found else "✗")  # type: ignore


def create_validation_report(output_dir: Path, toc_file: Path, spec_file: Path) -> Path:
    """Factory function to create validation report."""
    # Load data
    toc_data = []
    spec_data = []

    try:
        with open(toc_file, encoding="utf-8") as f:
            toc_data = [json.loads(line) for line in f if line.strip()]
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    try:
        with open(spec_file, encoding="utf-8") as f:
            spec_data = [json.loads(line) for line in f if line.strip()]
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    validator = XLSValidator(output_dir)  # Polymorphism
    return validator.generate_validation(toc_data, spec_data)
